"""XLSX revision-headers fixture generator.

Authors a baseline workbook with ``openpyxl``, then injects the
``xl/revisions/revisionHeaders.xml`` part (legacy shared-workbook
collaborative-edit metadata) into the zip alongside the required
``[Content_Types].xml`` registration and a relationship from
``xl/_rels/workbook.xml.rels``.

The on-disk shape matches what
``crates/xberg/src/extraction/excel.rs::parse_revision_headers_xml``
consumes: ``<header guid="{...}" userName="..." dateTime="..."/>`` under
``<headers xmlns="…spreadsheetml/2006/main">``.
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped, import-not-found, unused-ignore]

from .gt_schema import revisions_expectation, write_ground_truth

ZIP_MTIME = (2024, 1, 1, 0, 0, 0)

REV_HEADERS = [
    # (guid, userName, dateTime)
    ("11111111-1111-1111-1111-111111111111", "Alice", "2024-05-01T08:00:00Z"),
    ("22222222-2222-2222-2222-222222222222", "Bob", "2024-05-01T09:30:00Z"),
    ("33333333-3333-3333-3333-333333333333", "Carol", "2024-05-01T11:00:00Z"),
]

REVISION_HEADERS_RELID = "rIdRevHeaders"
REVISION_HEADERS_PATH = "xl/revisions/revisionHeaders.xml"
REVISION_HEADERS_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.revisionHeaders+xml"


def _build_baseline_xlsx() -> bytes:
    """Author a one-sheet workbook with three rows of data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Item"
    ws["B1"] = "Qty"
    ws["A2"] = "Widgets"
    ws["B2"] = 42
    ws["A3"] = "Gadgets"
    ws["B3"] = 7
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _revision_headers_xml() -> bytes:
    body = "".join(
        f'<header guid="{{{guid}}}" dateTime="{dt}" userName="{user}" maxSheetId="1"/>'
        for guid, user, dt in REV_HEADERS
    )
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<headers xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"{body}"
        "</headers>"
    )
    return xml.encode("utf-8")


def _patch_content_types(original: bytes) -> bytes:
    """Register the revisionHeaders content-type as an Override."""
    text = original.decode("utf-8")
    override = f'<Override PartName="/{REVISION_HEADERS_PATH}" ContentType="{REVISION_HEADERS_CT}"/>'
    if override in text:
        return original
    return text.replace("</Types>", f"{override}</Types>").encode("utf-8")


def _patch_workbook_rels(original: bytes) -> bytes:
    """Add a relationship from workbook -> revisionHeaders."""
    text = original.decode("utf-8")
    rel = (
        f'<Relationship Id="{REVISION_HEADERS_RELID}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/revisionHeaders" '
        'Target="revisions/revisionHeaders.xml"/>'
    )
    if REVISION_HEADERS_RELID in text:
        return original
    # ``</Relationships>`` should always be present; replace the last occurrence.
    return re.sub(r"</Relationships>\s*$", f"{rel}</Relationships>", text, count=1).encode("utf-8")


def _rewrite_zip(src_bytes: bytes, additions: dict[str, bytes], replacements: dict[str, bytes]) -> bytes:
    """Re-zip ``src_bytes`` with replacements applied and additions appended."""
    buf = io.BytesIO()
    seen: set[str] = set()
    with zipfile.ZipFile(io.BytesIO(src_bytes), "r") as src:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
            for name in src.namelist():
                data = replacements.get(name, src.read(name))
                info = zipfile.ZipInfo(name, ZIP_MTIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                dst.writestr(info, data)
                seen.add(name)
            for name, data in additions.items():
                if name in seen:
                    continue
                info = zipfile.ZipInfo(name, ZIP_MTIME)
                info.compress_type = zipfile.ZIP_DEFLATED
                dst.writestr(info, data)
    return buf.getvalue()


def generate(output_root: Path, repo_root: Path) -> list[Path]:
    """Emit xlsx_revisions_basic.xlsx + sidecar under ``output_root/xlsx/``."""
    output_dir = output_root / "xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    base = _build_baseline_xlsx()
    with zipfile.ZipFile(io.BytesIO(base), "r") as zf:
        content_types = zf.read("[Content_Types].xml")
        workbook_rels = zf.read("xl/_rels/workbook.xml.rels")

    out = _rewrite_zip(
        base,
        additions={REVISION_HEADERS_PATH: _revision_headers_xml()},
        replacements={
            "[Content_Types].xml": _patch_content_types(content_types),
            "xl/_rels/workbook.xml.rels": _patch_workbook_rels(workbook_rels),
        },
    )

    fixture_path = output_dir / "xlsx_revisions_basic.xlsx"
    sidecar_path = output_dir / "xlsx_revisions_basic.gt.json"
    fixture_path.write_bytes(out)
    write_ground_truth(
        sidecar_path,
        fixture_path,
        repo_root,
        document_format="xlsx",
        feature="revisions",
        expectations=revisions_expectation(
            expected_count=len(REV_HEADERS),
            revisions=[
                {
                    "kind": "FormatChange",
                    "author": user,
                    "timestamp": dt,
                    "revision_id": guid,
                }
                for guid, user, dt in REV_HEADERS
            ],
            notes=(
                "xl/revisions/revisionHeaders.xml carries shared-workbook collaborative-edit "
                "headers. The extractor maps each <header> to a DocumentRevision with kind = "
                "FormatChange (the closest neutral variant — header file does not record the "
                "kind of change). guid braces are stripped from revision_id."
            ),
        ),
        generator="xlsx_revisions",
    )
    return [fixture_path, sidecar_path]
